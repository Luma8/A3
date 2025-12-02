import os
from openpyxl import Workbook, load_workbook

FILE_NAME = "Formulário de Avaliação da Oficina(1-51) (1).xlsx"

def salvar_avaliacao_excel(dados: dict):
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Avaliações"

        ws.append(list(dados.keys()))
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append(list(dados.values()))
    wb.save(FILE_NAME)

    return True
